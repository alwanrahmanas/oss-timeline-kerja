# -*- coding: utf-8 -*-
"""
convert_to_js.py
================
Reads all Excel (.xlsx) timeline files from the `timeline-data/` folder and
converts them into `data_embed.js` for use in the Timeline Dashboard.

Excel Format Requirements
-------------------------
Row 1  : Month names (e.g. "Januari", "Februari", ..., "Desember")
Row 2  : Week labels  (e.g. "W1", "W2", "W3", ...)
Row 3+ : Activity data:
  - Col 2 : Program / Group name
  - Col 3 : Activity name (kegiatan)
  - Col 4 : Schedule text (optional)
  - Week columns (where Row 1/2 define the grid): put any marker to mark active weeks
           Accepted markers: ■  x  X  v  V  1  ✓  ●  ◆  ▪  •

Two week-label styles are supported:
  Style A (per-month) : W1..W5 resets every month → auto-detected by repetition
  Style B (global)    : W1..W53 unique across whole year

Team detection is done by filename (configure patterns in TIM_PATTERNS below).

Output
------
  data_embed.js   — JavaScript with TIMELINE_DATA and TIM_META constants.

Usage
-----
  python convert_to_js.py

Requirements
------------
  pip install openpyxl
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import json
import os
import re
import shutil
import tempfile
from datetime import datetime
from collections import Counter

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR  = os.path.join(BASE_DIR, "timeline-data")
OUTPUT_JS  = os.path.join(BASE_DIR, "data_embed.js")

# ── Team configuration ─────────────────────────────────────────────────────────
# Add your team filename patterns and metadata here.
# Patterns are matched case-insensitively against the Excel filename.
TIM_PATTERNS = {
    "team_a": r"team.?a|team_a",
    "team_b": r"team.?b|team_b",
    "team_c": r"team.?c|team_c",
}

TIM_META = {
    "team_a": {"label": "Team A", "color": "#01696f"},
    "team_b": {"label": "Team B", "color": "#006494"},
    "team_c": {"label": "Team C", "color": "#7a39bb"},
}

# ── Year/Calendar configuration ────────────────────────────────────────────────
# List of (month_number, display_name, num_weeks_in_month)
# Total weeks must sum to 53 for a standard year.
MONTHS_CONFIG = [
    (1,  "Januari",   5),
    (2,  "Februari",  4),
    (3,  "Maret",     4),
    (4,  "April",     5),
    (5,  "Mei",       4),
    (6,  "Juni",      4),
    (7,  "Juli",      5),
    (8,  "Agustus",   4),
    (9,  "September", 4),
    (10, "Oktober",   5),
    (11, "November",  4),
    (12, "Desember",  5),
]

# ── Internal helpers ───────────────────────────────────────────────────────────
def build_week_to_month():
    w2m = {}
    w = 1
    for mnum, _, nweeks in MONTHS_CONFIG:
        for _ in range(nweeks):
            w2m[w] = mnum
            w += 1
    return w2m


WEEK_TO_MONTH = build_week_to_month()


def build_month_col_offsets(row1, row2, ws, max_col):
    """
    Build a mapping {column_index → global_week_number (1-53)}.
    Handles both Style A (per-month labels) and Style B (global labels).
    Robust against typos in week cell labels (uses positional index within month span).
    """
    # Collect month header positions from row 1
    month_spans = []
    for c in range(1, max_col + 1):
        v = ws.cell(row1, c).value
        if v and isinstance(v, str):
            for mnum, mname, _ in MONTHS_CONFIG:
                if mname.lower() in v.lower():
                    month_spans.append((c, mnum))

    # Collect week label columns from row 2
    week_labels = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row2, c).value
        if v and isinstance(v, str) and re.match(r'^W\d+$', v.strip(), re.IGNORECASE):
            week_labels[c] = int(v.strip()[1:])

    if not week_labels:
        return {}

    # Detect style: if W1 appears multiple times → per-month (Style A)
    wn_counts = Counter(week_labels.values())
    has_repeats = any(cnt > 1 for cnt in wn_counts.values())
    is_per_month = has_repeats or (len(month_spans) >= 6 and wn_counts.get(1, 0) >= 6)

    if not is_per_month:
        # Style B: use labels directly as global week numbers
        return {col: wn for col, wn in week_labels.items()}

    # Style A: positional index within each month span
    col_to_global_week = {}
    month_spans_sorted = sorted(month_spans, key=lambda x: x[0])

    month_cumulative = {}
    cumul = 1
    for mnum, _, nweeks in MONTHS_CONFIG:
        month_cumulative[mnum] = cumul
        cumul += nweeks

    for i, (mcol, mnum) in enumerate(month_spans_sorted):
        next_mcol = month_spans_sorted[i + 1][0] if i + 1 < len(month_spans_sorted) else max_col + 1
        base_week = month_cumulative.get(mnum, 1)
        month_week_cols = sorted(wc for wc in week_labels if mcol <= wc < next_mcol)
        for idx, wc in enumerate(month_week_cols):
            col_to_global_week[wc] = min(base_week + idx, 53)

    return col_to_global_week


def detect_tim_key(filename, ws):
    """Detect team key from filename. Falls back to cleaned filename."""
    fname_lower = filename.lower()
    for key, pattern in TIM_PATTERNS.items():
        if re.search(pattern, fname_lower, re.IGNORECASE):
            return key
    # Fallback: use cleaned filename as key
    clean = re.sub(r'[_\-\s]+', '_', re.sub(r'\.(xlsx?)$', '', filename, flags=re.IGNORECASE)).lower()
    return clean


def is_marker(v):
    """Return True if a cell value represents an active schedule marker."""
    if v is None:
        return False
    if isinstance(v, (int, float)) and v == 1:
        return True
    return str(v).strip() in {"■", "x", "X", "v", "V", "1", "✓", "●", "◆", "▪", "•"}


def load_workbook_safe(fpath, fname):
    """Load workbook, falling back to a temp copy if the file is locked."""
    try:
        return openpyxl.load_workbook(fpath, data_only=True)
    except PermissionError:
        tmp_path = os.path.join(tempfile.gettempdir(), fname)
        try:
            shutil.copy2(fpath, tmp_path)
            print(" (copied from locked file)", end="")
            return openpyxl.load_workbook(tmp_path, data_only=True)
        except Exception as e:
            print(f" ERROR: locked and copy failed — {e}")
            return None
    except Exception as e:
        print(f" ERROR: {e}")
        return None


def parse_sheet(ws, tim_key, id_offset):
    """Parse a single worksheet and return a list of activity dicts."""
    max_row, max_col = ws.max_row, ws.max_column
    header_row1, header_row2 = 1, 2

    # Find first data row (usually row 3, sometimes 4)
    data_start_row = 3
    for r in range(3, min(8, max_row + 1)):
        if ws.cell(r, 2).value or ws.cell(r, 3).value:
            data_start_row = r
            break

    col_to_week = build_month_col_offsets(header_row1, header_row2, ws, max_col)
    if not col_to_week:
        print(f"  WARNING: no week columns found in sheet '{ws.title}'")
        return []

    week_cols = sorted(col_to_week.keys())
    results = []
    uid = id_offset

    for r in range(data_start_row, max_row + 1):
        program  = ws.cell(r, 2).value
        kegiatan = ws.cell(r, 3).value

        if not kegiatan or not isinstance(kegiatan, str) or not kegiatan.strip():
            continue
        if not program or not isinstance(program, str) or not program.strip():
            program = "(No Program)"

        active_weeks = sorted(set(
            col_to_week[wc]
            for wc in week_cols
            if is_marker(ws.cell(r, wc).value) and 1 <= col_to_week[wc] <= 53
        ))

        if not active_weeks:
            continue

        jadwal_raw = ws.cell(r, 4).value
        selesai_raw = ws.cell(r, 5).value if max_col >= 5 else None

        if jadwal_raw and isinstance(jadwal_raw, str) and jadwal_raw.strip():
            jadwal_teks = jadwal_raw.strip()
        elif jadwal_raw and selesai_raw:
            jadwal_teks = f"{jadwal_raw} – {selesai_raw}"
        else:
            jadwal_teks = f"W{min(active_weeks)} – W{max(active_weeks)}"

        active_months = sorted(set(WEEK_TO_MONTH.get(w, 1) for w in active_weeks))

        results.append({
            "id":           uid,
            "tim":          tim_key,
            "program":      program.strip(),
            "kegiatan":     kegiatan.strip(),
            "jadwal_teks":  jadwal_teks,
            "minggu_aktif": active_weeks,
            "bulan_aktif":  active_months,
        })
        uid += 1

    return results


def convert():
    if not os.path.isdir(EXCEL_DIR):
        print(f"ERROR: Folder '{EXCEL_DIR}' not found.")
        print("  Please create a 'timeline-data/' folder and put your .xlsx files there.")
        return

    excel_files = sorted(f for f in os.listdir(EXCEL_DIR) if f.lower().endswith(".xlsx"))
    if not excel_files:
        print("ERROR: No .xlsx files found in 'timeline-data/'")
        return

    print(f"Found {len(excel_files)} file(s):")
    by_tim = {}
    tim_order = list(TIM_PATTERNS.keys())
    id_counter = 0

    for fname in excel_files:
        fpath = os.path.join(EXCEL_DIR, fname)
        print(f"  [{fname}]", end="")

        wb = load_workbook_safe(fpath, fname)
        if wb is None:
            continue

        ws = wb.worksheets[0]
        tim_key = detect_tim_key(fname, ws)

        if tim_key not in by_tim:
            by_tim[tim_key] = []
        if tim_key not in tim_order:
            tim_order.append(tim_key)
        if tim_key not in TIM_META:
            TIM_META[tim_key] = {"label": tim_key.replace("_", " ").title(), "color": "#888888"}

        rows = parse_sheet(ws, tim_key, id_counter)
        by_tim[tim_key].extend(rows)
        id_counter += len(rows)
        print(f" → {len(rows)} activities [{tim_key}]")
        wb.close()

    # Reassign sequential IDs in team order
    all_data = []
    uid = 0
    for tk in tim_order:
        for item in by_tim.get(tk, []):
            item["id"] = uid
            uid += 1
            all_data.append(item)

    tim_meta_out = {k: TIM_META[k] for k in tim_order if by_tim.get(k)}
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    js_content = f"""// AUTO-GENERATED by convert_to_js.py — {generated_at}
// Do not edit manually. Run `python convert_to_js.py` to regenerate.

const GENERATED_AT = "{generated_at}";

const TIM_META = {json.dumps(tim_meta_out, ensure_ascii=False, indent=2)};

const TIMELINE_DATA = {json.dumps(all_data, ensure_ascii=False, indent=2)};
"""

    with open(OUTPUT_JS, "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f"\nDone! {uid} activities from {len(tim_meta_out)} team(s) written to {OUTPUT_JS}")
    for tk in tim_order:
        if by_tim.get(tk):
            print(f"  {tk:20s}: {len(by_tim[tk])} activities")


if __name__ == "__main__":
    convert()
